"""
Parse GTAOL全街车稀有色大全.xlsx
Outputs street-car-colors.json with:
  - 主色调 (primary), 副色调 (secondary), 珠光 (pearlescent), 轮毂色 (wheel)
  - Rare cells identified by RED font (FFFF0000)
  - Color IDs resolved against gta-colors.json for hex values
"""
import openpyxl
import json
import re
import sys

SRC = r'C:\Users\lenovo\Downloads\GTAOL全街车稀有色大全.xlsx'
COLORS_JSON = r'c:\Users\lenovo\Desktop\ai\new gta\gta\gta-colors.json'
OUTPUT = r'c:\Users\lenovo\Desktop\ai\new gta\gta\street-car-colors.json'

# Load color lookup
with open(COLORS_JSON, 'r', encoding='utf-8') as f:
    color_by_id = {c['id']: c for c in json.load(f)['colors']}

warned_ids = set()

# Pattern to match "ID: Name" or "ID：Name" (both regular and full-width colon)
COLOR_LINE_RE = re.compile(r'^(\d+)[:：]\s*(.+)')

def parse_color_cell(text, font_rgb):
    """Parse '73: Racing Blue' into structured object."""
    is_rare = (font_rgb and 'FFFF0000' in str(font_rgb))
    if not text:
        return None
    text = str(text).strip()
    m = COLOR_LINE_RE.match(text)
    if not m:
        return None
    cid = int(m.group(1))
    cname = m.group(2).strip()
    hex_val = ''
    name_cn = ''
    price = 0
    unlock = ''
    if cid in color_by_id:
        hex_val = color_by_id[cid]['hex']
        name_cn = color_by_id[cid]['name_cn']
        price = color_by_id[cid].get('price', 0)
        unlock = color_by_id[cid].get('unlock', '')
    elif cid not in warned_ids:
        warned_ids.add(cid)
        print(f'  [WARN] Color ID {cid} not found in gta-colors.json')
    return {
        'id': cid,
        'name': cname,
        'name_cn': name_cn,
        'hex': hex_val,
        'price': price,
        'unlock': unlock,
        'isRare': is_rare
    }


wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb.active

vehicles = []
i = 1  # 1-indexed row number

while i <= ws.max_row:
    cell_a = ws.cell(row=i, column=1).value
    cell_b = ws.cell(row=i, column=2).value

    # Detect model name row: col 1 contains "模型名称"
    if cell_a and '模型名称' in str(cell_a):
        model_name = str(cell_b).strip().upper() if cell_b else ''
        if not model_name:
            i += 1
            continue

        # Get vehicle display name (usually 2 rows above, but can be 1 row above)
        name_cn = ''
        if i >= 2:
            name_cn = str(ws.cell(row=i - 2, column=1).value or '').strip()

        # Get brand (row above, col 2)
        brand = ''
        if i >= 1:
            brand = str(ws.cell(row=i - 1, column=2).value or '').strip()

        # Get DLC (row after model, col 2)
        dlc = ''
        if i + 1 <= ws.max_row:
            d0 = str(ws.cell(row=i + 1, column=1).value or '')
            if 'DLC:' in d0:
                dlc = str(ws.cell(row=i + 1, column=2).value or '').strip()

        # Get vehicle class (row after DLC, col 2)
        vclass = ''
        if i + 2 <= ws.max_row:
            v0 = str(ws.cell(row=i + 2, column=1).value or '')
            if '用途' in v0 or '用' in v0:
                vclass = str(ws.cell(row=i + 2, column=2).value or '').strip()

        # Read color rows after header row
        color_rows = []
        all_rare = False
        no_special = False
        j = i + 4

        while j <= ws.max_row:
            # Check if all 4 cells are empty → end of this vehicle
            all_empty = all(
                ws.cell(row=j, column=c).value is None
                for c in range(1, 5)
            )
            if all_empty:
                break

            # Check if next vehicle started
            c1 = str(ws.cell(row=j, column=1).value or '')
            if '模型名称' in c1:
                break

            # Check for "所有颜色都是稀有颜色"
            if '所有颜色' in c1 and '稀有' in c1:
                all_rare = True
                j += 1
                break

            # Check for "无特殊色" — no rare colors for this vehicle
            if '无特殊色' in c1:
                no_special = True
                j += 1
                break

            # Check for "无法直接获得" — can't obtain from street
            if '无法直接获得' in c1 or '无法' in c1:
                no_special = True
                j += 1
                break

            # Try to parse color data row
            c0_val = ws.cell(row=j, column=1).value
            if c0_val and isinstance(c0_val, str):
                m = COLOR_LINE_RE.match(c0_val.strip())
                if m:
                    # Parse all 4 columns
                    row_data = {}
                    for col_idx, key in enumerate(['primary', 'secondary', 'pearlescent', 'wheel']):
                        cell = ws.cell(row=j, column=col_idx + 1)
                        font_rgb = cell.font.color.rgb if cell.font and cell.font.color else None
                        row_data[key] = parse_color_cell(cell.value, font_rgb)
                    color_rows.append(row_data)
                    j += 1
                    continue

            # If we get here, row is not empty but also not a recognized pattern
            # Could be a model identifier row ("F620", "BMX", etc.) — just skip
            j += 1

        if model_name:
            vehicles.append({
                'model_name': model_name,
                'name_cn': name_cn,
                'brand': brand,
                'dlc': dlc,
                'vehicle_class': vclass,
                'all_rare': all_rare,
                'no_special': no_special,
                'color_rows': color_rows
            })

    i += 1

# Save
with open(OUTPUT, 'w', encoding='utf-8') as f:
    json.dump(vehicles, f, ensure_ascii=False, indent=2)

# Stats
with_colors = sum(1 for v in vehicles if v['color_rows'])
all_rare = sum(1 for v in vehicles if v['all_rare'])
no_special = sum(1 for v in vehicles if v['no_special'])
rare_cells = sum(
    sum(1 for cell in row.values() if cell and cell.get('isRare'))
    for v in vehicles for row in v['color_rows']
)

print(f'Vehicles: {len(vehicles)}')
print(f'  With color rows: {with_colors}')
print(f'  All rare: {all_rare}')
print(f'  No special: {no_special}')
print(f'  Rare cells marked: {rare_cells}')
